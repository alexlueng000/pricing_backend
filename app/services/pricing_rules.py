import csv
from datetime import date
from io import BytesIO, StringIO

from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.models.pricing_rule import PricingRule
from app.repositories.pricing_rules import PricingRuleRepository
from app.schemas.pricing_rule import PricingRuleCreate


class PricingRuleService:
    def __init__(self, db: Session) -> None:
        self.db = db
        self.rules = PricingRuleRepository(db)

    def create_rule(self, payload: PricingRuleCreate) -> PricingRule:
        rule = PricingRule(**payload.model_dump())
        self.rules.add(rule)
        self.db.commit()
        self.db.refresh(rule)
        return rule

    def import_rules(
        self,
        *,
        filename: str,
        content: bytes,
    ) -> tuple[int, int]:
        rows = self._read_import_rows(filename=filename, content=content)
        imported_count = 0
        skipped_count = 0

        for row in rows:
            normalized = self._normalize_import_row(row)
            if not self._has_required_import_fields(normalized):
                skipped_count += 1
                continue
            rule = PricingRule(**normalized)
            self.rules.add(rule)
            imported_count += 1

        if imported_count:
            self.db.commit()
        return imported_count, skipped_count

    def list_active_rules(
        self,
        *,
        country_region: str | None = None,
        patent_type: str | None = None,
        filing_route: str | None = None,
        currency: str | None = None,
        status: str = "active",
    ) -> list[PricingRule]:
        return self.rules.list_active(
            country_region=country_region,
            patent_type=patent_type,
            filing_route=filing_route,
            currency=currency,
            status=status,
        )

    def find_quote_rules(
        self,
        *,
        country_region: str,
        patent_type: str,
        filing_route: str,
        entity_type: str | None,
        quote_date: date,
    ) -> list[PricingRule]:
        return self.rules.find_active_rules(
            country_region=country_region,
            patent_type=patent_type,
            filing_route=filing_route,
            entity_type=entity_type,
            quote_date=quote_date,
        )

    def _read_import_rows(self, *, filename: str, content: bytes) -> list[dict[str, object]]:
        lowered = filename.lower()
        if lowered.endswith(".csv"):
            text = content.decode("utf-8-sig")
            return list(csv.DictReader(StringIO(text)))
        if lowered.endswith(".xlsx"):
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            worksheet = workbook.active
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(value).strip() if value is not None else "" for value in rows[0]]
            return [
                {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
                for row in rows[1:]
            ]
        raise ValueError("仅支持 .csv 或 .xlsx 规则导入文件")

    def _normalize_import_row(self, row: dict[str, object]) -> dict[str, object]:
        header_map = {
            "国家/地区": "country_region",
            "国家": "country_region",
            "country": "country_region",
            "country_region": "country_region",
            "专利类型": "patent_type",
            "patent_type": "patent_type",
            "申请路径": "filing_route",
            "进入路径": "filing_route",
            "filing_route": "filing_route",
            "实体类型": "entity_type",
            "entity_type": "entity_type",
            "费用阶段": "fee_stage",
            "阶段": "fee_stage",
            "fee_stage": "fee_stage",
            "费用项目": "fee_item",
            "项目": "fee_item",
            "fee_item": "fee_item",
            "币种": "currency",
            "currency": "currency",
            "官费公式": "official_fee_formula",
            "official_fee_formula": "official_fee_formula",
            "外代费公式": "foreign_agent_fee_formula",
            "外所费公式": "foreign_agent_fee_formula",
            "foreign_agent_fee_formula": "foreign_agent_fee_formula",
            "本所代理费公式": "local_agent_fee_formula",
            "local_agent_fee_formula": "local_agent_fee_formula",
            "开票税费处理方式": "invoice_tax_policy",
            "invoice_tax_policy": "invoice_tax_policy",
            "条件表达式": "condition_expression",
            "condition_expression": "condition_expression",
            "客户备注": "customer_remark",
            "customer_remark": "customer_remark",
            "内部备注": "internal_note",
            "internal_note": "internal_note",
            "生效日期": "effective_date",
            "effective_date": "effective_date",
            "状态": "status",
            "status": "status",
        }
        normalized: dict[str, object] = {}
        for key, value in row.items():
            mapped_key = header_map.get(str(key).strip())
            if mapped_key is None:
                continue
            normalized[mapped_key] = self._clean_import_value(
                value,
                keep_native=mapped_key == "effective_date",
            )
        if normalized.get("effective_date") is not None:
            normalized["effective_date"] = self._parse_import_date(
                normalized["effective_date"]
            )
        normalized.setdefault("invoice_tax_policy", "tax_included")
        normalized.setdefault("status", "active")
        return normalized

    @staticmethod
    def _clean_import_value(value: object, *, keep_native: bool = False) -> object | None:
        if value is None:
            return None
        if isinstance(value, str):
            stripped = value.strip()
            return stripped or None
        if keep_native:
            return value
        text = str(value).strip()
        return text or None

    @staticmethod
    def _has_required_import_fields(row: dict[str, object]) -> bool:
        required_fields = {
            "country_region",
            "patent_type",
            "filing_route",
            "fee_stage",
            "fee_item",
            "currency",
            "effective_date",
        }
        return all(row.get(field) for field in required_fields)

    @staticmethod
    def _parse_import_date(value: object) -> date:
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            try:
                return date.fromisoformat(value.strip())
            except ValueError as exc:
                raise ValueError("生效日期需使用 yyyy-mm-dd 格式") from exc
        raise ValueError("生效日期需使用 yyyy-mm-dd 格式")
